from flask import Flask, render_template, request, redirect, url_for, session, send_file, abort, flash, send_from_directory
from flask_socketio import SocketIO, emit, join_room, leave_room
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime
from openpyxl import Workbook
from io import BytesIO
from flask import flash, redirect, url_for
import sqlite3
import os
import pytz





BASE_DIR = os.path.abspath(os.path.dirname(__file__))
DB_PATH = os.path.join(BASE_DIR, "app.db")
UPLOAD_REPORT = os.path.join("static", "uploads", "reports")
UPLOAD_CERT = os.path.join("static", "uploads", "certificates")
ALLOWED_CERT_EXT = {".pdf", ".jpg", ".jpeg", ".png"}

app = Flask(__name__)
socketio = SocketIO(app, cors_allowed_origins="*")
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY","supersecret-session")
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024  # 32 MB
UPLOAD_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'uploads')

if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

os.makedirs(UPLOAD_REPORT, exist_ok=True)
os.makedirs(UPLOAD_CERT, exist_ok=True)

BRAND = {
    "primary": "#12B886",
    "secondary": "#0CA678",
    "dark": "#0B5C49"
}


@app.route('/uploads/<path:filename>')
def uploaded_file(filename):
    """Tampilkan file pemeliharaan, sertifikat, report, dll."""
    return send_from_directory(os.path.join(app.root_path, "static", "uploads"), filename)



@app.context_processor
def inject_user():
    if 'user' in session:
        return {'user': session['user']}
    return {'user': None}


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def ensure_tables():
    with db() as conn:
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS users(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('admin','user'))
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS reports(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            unit TEXT NOT NULL,
            description TEXT NOT NULL,
            image_path TEXT,
            status TEXT NOT NULL DEFAULT 'Proses',
            comment TEXT,
            created_at TEXT NOT NULL,
            user_id INTEGER,
            FOREIGN KEY(user_id) REFERENCES users(id)
        )''')
        c.execute('''CREATE TABLE IF NOT EXISTS inventory(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            alat_nama TEXT,
            alat_merk TEXT,
            alat_tipe TEXT,
            alat_sn TEXT,
            ruangan TEXT,
            tahun_perolehan INTEGER,
            sumber_dana TEXT,
            tgl_kalibrasi TEXT,
            tgl_jatuh_tempo TEXT,
            sertifikat_path TEXT
        )''')
        
        # ✅ Tabel PEMELIHARAAN alat kesehatan
        c.execute('''CREATE TABLE IF NOT EXISTS pemeliharaan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nama_alat TEXT NOT NULL,
            merk TEXT,
            tipe TEXT,
            sn TEXT,
            ruangan TEXT,
            tgl_pemeliharaan TEXT,
            kondisi TEXT,
            file_path TEXT
        )''')

        # 🆕 Tabel SPO
        c.execute('''CREATE TABLE IF NOT EXISTS spo(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            alat_nama TEXT NOT NULL,
            alat_merk TEXT,
            alat_tipe TEXT,
            file_path TEXT
        )''')
        
        # 🧩 Tambah tabel stock barang
        c.execute('''CREATE TABLE IF NOT EXISTS stock_in (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nama_barang TEXT NOT NULL,
            qty INTEGER NOT NULL,
            harga REAL NOT NULL,
            tgl_pembelian TEXT NOT NULL
        )''')

        c.execute('''CREATE TABLE IF NOT EXISTS stock_out (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nama_barang TEXT NOT NULL,
            qty INTEGER NOT NULL,
            ruangan TEXT,
            penanggung_jawab TEXT,
            tgl_keluar TEXT NOT NULL,
            bukti_path TEXT
        )''')


       
        # ✅ Tambahkan index untuk optimasi query
        c.execute("CREATE INDEX IF NOT EXISTS idx_reports_user_id ON reports(user_id)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_reports_created_at ON reports(created_at)")

        # 🗨️ Tambah tabel chat
        c.execute('''CREATE TABLE IF NOT EXISTS messages(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sender_id INTEGER NOT NULL,
            receiver_id INTEGER NOT NULL,
            message TEXT NOT NULL,
            created_at TEXT NOT NULL,
            FOREIGN KEY(sender_id) REFERENCES users(id),
            FOREIGN KEY(receiver_id) REFERENCES users(id)
        )''')
        conn.commit()
        

def ensure_admin():
    with db() as conn:
        c = conn.cursor()
        c.execute("SELECT id FROM users WHERE role='admin' LIMIT 1")
        row = c.fetchone()
        if not row:
            username = "admin"
            import secrets, string
            pwd = os.environ.get("ADMIN_PASSWORD","".join(secrets.choice(string.ascii_letters+string.digits) for _ in range(10)))
            c.execute("INSERT INTO users(username,password_hash,role) VALUES(?,?,?)",
                      (username, generate_password_hash(pwd), "admin"))
            conn.commit()
            print("=== ADMIN AKUN DIBUAT ===")
            print("Username: admin")
            print("Password:", pwd)
            print("Silakan login dan tambahkan user lain. (Password TIDAK ditampilkan di halaman login)")

def login_required(role=None):
    def decorator(fn):
        from functools import wraps
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not session.get("user"):
                return redirect(url_for("login"))
            if role and session["user"]["role"] != role:
                return abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return decorator

@app.route("/")
def index():
    if not session.get("user"):
        return redirect(url_for("login"))
    return redirect(url_for("admin_dashboard" if session["user"]["role"]=="admin" else "user_dashboard"))

@app.route("/login", methods=["GET","POST"])
def login():
    err = None
    if request.method == "POST":
        username = request.form.get("username","").strip()
        password = request.form.get("password","")
        with db() as conn:
            c = conn.cursor()
            c.execute("SELECT * FROM users WHERE username=?", (username,))
            user = c.fetchone()
        if not user or not check_password_hash(user["password_hash"], password):
            err = "Username atau password salah"
        else:
            session["user"] = {"id": user["id"], "username": user["username"], "role": user["role"]}
            return redirect(url_for("index"))
    return render_template("login.html", error=err, BRAND=BRAND)

@app.route("/logout")
def logout():
    global joined_users
    if "user" in session:
        joined_users.discard(session["user"]["username"])  # hapus dari daftar
    session.clear()
    return redirect(url_for("login"))


# ---------- ADMIN ----------
@app.route("/admin")
def admin_dashboard():
    if not session.get("user") or session["user"]["role"]!="admin":
        return redirect(url_for("login"))

    q = (request.args.get("q") or "").strip()
    status_filter = (request.args.get("status_filter") or "").strip()
    page = int(request.args.get("page", 1))
    per_page = 50
    offset = (page - 1) * per_page

    sql = "SELECT * FROM reports WHERE 1=1"
    params = []

    if q:
        sql += " AND (title LIKE ? OR unit LIKE ?)"
        keyword = f"%{q}%"
        params.extend([keyword, keyword])

    if status_filter:
        sql += " AND status = ?"
        params.append(status_filter)

    sql += " ORDER BY datetime(created_at) DESC LIMIT ? OFFSET ?"
    params.extend([per_page, offset])

    with db() as conn:
        reports = conn.execute(sql, params).fetchall()
        inv_total = conn.execute("SELECT COUNT(*) as total FROM inventory").fetchone()["total"]
        rep_total = conn.execute("SELECT COUNT(*) as total FROM reports").fetchone()["total"]
        total_reports = conn.execute("SELECT COUNT(*) as cnt FROM reports").fetchone()["cnt"]
        total_pages = (total_reports + per_page - 1) // per_page
        
         # 💡 Tambahan: hitung jumlah laporan berdasarkan status
        total_proses = conn.execute(
            "SELECT COUNT(*) as cnt FROM reports WHERE status='Proses'"
        ).fetchone()["cnt"]
        total_selesai = conn.execute(
            "SELECT COUNT(*) as cnt FROM reports WHERE status='Selesai'"
        ).fetchone()["cnt"]
        
         # 📊 Hitung laporan per bulan (dibagi status)
        monthly_stats = conn.execute("""
            SELECT 
                strftime('%Y-%m', created_at) AS bulan,
                SUM(CASE WHEN status='Proses' THEN 1 ELSE 0 END) AS proses,
                SUM(CASE WHEN status='Selesai' THEN 1 ELSE 0 END) AS selesai
            FROM reports
            GROUP BY bulan
            ORDER BY bulan ASC
        """).fetchall()

        # ubah ke list untuk dikirim ke template
        bulan_labels = [row["bulan"] for row in monthly_stats]
        proses_data = [row["proses"] for row in monthly_stats]
        selesai_data = [row["selesai"] for row in monthly_stats]
        
    return render_template("admin_dashboard.html",
                           reports=reports,
                           invTotal=inv_total,
                           repTotal=rep_total,
                           total_proses=total_proses,   # 💡 dikirim ke HTML
                           total_selesai=total_selesai, # 💡 dikirim ke HTML
                           bulan_labels=bulan_labels,    # ✅ kirim ke template
                           proses_data=proses_data,
                           selesai_data=selesai_data,
                           BRAND=BRAND,
                           user=session.get("user"),
                           q=q, status_filter=status_filter,
                           page=page, total_pages=total_pages,
                           per_page=per_page)


@app.post("/admin/report/<int:rid>/update")
def admin_report_update(rid):
    if not session.get("user") or session["user"]["role"]!="admin":
        return abort(403)
    status = request.form.get("status", "Proses")
    comment = request.form.get("comment","").strip()

    import pytz
    tz = pytz.timezone("Asia/Jakarta")
    updated_at = datetime.now(tz).strftime("%Y-%m-%d %H:%M:%S")

    with db() as conn:
        conn.execute(
            "UPDATE reports SET status=?, comment=?, updated_at=? WHERE id=?",
            (status, comment, updated_at, rid)
        )
        conn.commit()
    return redirect(url_for("admin_dashboard"))


@app.post("/admin/report/<int:rid>/delete")
def admin_report_delete(rid):
    if not session.get("user") or session["user"]["role"]!="admin":
        return abort(403)
    with db() as conn:
        c = conn.cursor()
        c.execute("SELECT image_path FROM reports WHERE id=?", (rid,))
        row = c.fetchone()
        if row and row["image_path"]:
            try: os.remove(os.path.join(BASE_DIR, row["image_path"]))
            except: pass
        conn.execute("DELETE FROM reports WHERE id=?", (rid,))
        conn.commit()
    return redirect(url_for("admin_dashboard"))
    
# ---------- EXPORT REPORTS (ADMIN ONLY) ----------
@app.get("/admin/reports/export/csv")
def admin_reports_export_csv():
    if not session.get("user") or session["user"]["role"]!="admin":
        return abort(403)
    with db() as conn:
        rows = conn.execute("SELECT * FROM reports ORDER BY datetime(created_at) DESC").fetchall()
    headers = ["Tanggal", "Judul", "Unit", "Deskripsi", "Status", "Komentar", "User ID"]
    lines = [",".join(headers)]
    for r in rows:
        vals = [r["created_at"], r["title"], r["unit"], r["description"], r["status"], r["comment"] or "", str(r["user_id"])]
        esc = ['"{}"'.format((str(v) if v else "").replace('"','""')) for v in vals]
        lines.append(",".join(esc))
    bio = BytesIO()
    bio.write("\n".join(lines).encode("utf-8"))
    bio.seek(0)
    return send_file(bio, mimetype="text/csv", as_attachment=True, download_name="laporan.csv")


@app.get("/admin/reports/export/xlsx")
def admin_reports_export_xlsx():
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)
    
    # 🔹 Ganti query agar JOIN ke tabel users
    with db() as conn:
        rows = conn.execute("""
            SELECT 
                r.created_at,
                r.title,
                r.unit,
                r.description,
                r.status,
                r.comment,
                u.username AS username
            FROM reports AS r
            LEFT JOIN users AS u ON r.user_id = u.id
            ORDER BY datetime(r.created_at) DESC
        """).fetchall()

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Alat Kesehatan"

    # 🏷️ Header kolom
    headers = ["Tanggal", "Judul", "Unit/Ruangan", "Deskripsi", "Status", "Komentar", "Username"]
    ws.append(headers)

    # 🎨 Styling header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="12B886", end_color="12B886", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # 📊 Data rows
    for r in rows:
        ws.append([
            r["created_at"],
            r["title"],
            r["unit"],
            r["description"],
            r["status"],
            r["comment"] or "",
            r["username"] or "-",  # ✅ tampilkan nama user, bukan angka
        ])

    # 📏 Format seluruh sel
    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

    # ⚙️ Auto width
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    ws.row_dimensions[1].height = 25

    # 💾 Simpan dan kirim
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="laporan_kerusakan.xlsx"
    )



@app.get("/admin/users")
def admin_users():
    if not session.get("user") or session["user"]["role"]!="admin":
        return abort(403)
    with db() as conn:
        users = conn.execute("SELECT id,username,role FROM users ORDER BY id DESC").fetchall()
    return render_template("admin_users.html", users=users, BRAND=BRAND, user=session.get("user"))

@app.post("/admin/users/add")
def admin_users_add():
    if not session.get("user") or session["user"]["role"]!="admin":
        return abort(403)
    username = request.form.get("username","").strip()
    password = request.form.get("password","")
    role = request.form.get("role","user")
    if not username or not password: return redirect(url_for("admin_users"))
    with db() as conn:
        try:
            conn.execute("INSERT INTO users(username,password_hash,role) VALUES(?,?,?)",
                         (username, generate_password_hash(password), role))
            conn.commit()
        except sqlite3.IntegrityError:
            pass
    return redirect(url_for("admin_users"))
    
@app.post("/admin/users/<int:uid>/delete")
def admin_users_delete(uid):
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)
    with db() as conn:
        conn.execute("DELETE FROM users WHERE id=?", (uid,))
        conn.commit()
    return redirect(url_for("admin_users"))

# ---------- INVENTORY (ADMIN ONLY) ----------
@app.get("/admin/inventory")
def admin_inventory():
    if not session.get("user") or session["user"]["role"]!="admin":
        return abort(403)

    # Ambil parameter search & filter dari query string
    q = (request.args.get("q") or "").strip()
    filter_ruangan = (request.args.get("filter_ruangan") or "").strip()
    filter_tahun = (request.args.get("filter_tahun") or "").strip()

    # Pagination
    page = int(request.args.get("page", 1))
    per_page = 50
    offset = (page - 1) * per_page

    # Query dasar
    sql = "SELECT * FROM inventory WHERE 1=1"
    params = []

    # Search keyword
    if q:
        like = f"%{q}%"
        sql += " AND (alat_nama LIKE ? OR alat_merk LIKE ? OR alat_tipe LIKE ? OR alat_sn LIKE ? OR ruangan LIKE ? OR sumber_dana LIKE ?)"
        params.extend([like]*6)

    # Filter ruangan
    if filter_ruangan:
        sql += " AND ruangan LIKE ?"
        params.append(f"%{filter_ruangan}%")


    # Filter tahun
    if filter_tahun:
        sql += " AND tahun_perolehan=?"
        params.append(filter_tahun)

    sql += " ORDER BY id DESC LIMIT ? OFFSET ?"
    params.extend([per_page, offset])

    with db() as conn:
        items = conn.execute(sql, params).fetchall()
        total_items = conn.execute("SELECT COUNT(*) as cnt FROM inventory").fetchone()["cnt"]
        total_pages = (total_items + per_page - 1) // per_page

    return render_template("admin_inventory.html", 
                           items=items, 
                           q=q, 
                           filter_ruangan=filter_ruangan,
                           filter_tahun=filter_tahun,
                           BRAND=BRAND, 
                           user=session.get("user"),
                           page=page,
                           total_pages=total_pages,
                           per_page=per_page)

@app.post("/admin/inventory/<int:iid>/delete")
def admin_inventory_delete(iid):
    if not session.get("user") or session["user"]["role"]!="admin":
        return abort(403)
    with db() as conn:
        row = conn.execute("SELECT sertifikat_path FROM inventory WHERE id=?", (iid,)).fetchone()
        if row and row["sertifikat_path"]:
            try: os.remove(os.path.join(BASE_DIR, row["sertifikat_path"]))
            except: pass
        conn.execute("DELETE FROM inventory WHERE id=?", (iid,))
        conn.commit()
    return redirect(url_for("admin_inventory"))

@app.get("/admin/inventory/export/csv")
def admin_inventory_export_csv():
    if not session.get("user") or session["user"]["role"]!="admin":
        return abort(403)
    with db() as conn:
        rows = conn.execute("SELECT * FROM inventory ORDER BY id DESC").fetchall()
    headers = ["Nama Alat","Merk Alat","Type Alat","SN Alat","Ruangan","Tahun Perolehan","Sumber Dana","Tanggal Kalibrasi","Tanggal Jatuh Tempo","Path Sertifikat"]
    lines = [",".join(headers)]
    for r in rows:
        vals = [r["alat_nama"], r["alat_merk"], r["alat_tipe"], r["alat_sn"], r["ruangan"],
                r["tahun_perolehan"], r["sumber_dana"], r["tgl_kalibrasi"], r["tgl_jatuh_tempo"], r["sertifikat_path"] or ""]
        esc = ['"{}"'.format((str(v) if v is not None else "").replace('"','""')) for v in vals]
        lines.append(",".join(esc))
    bio = BytesIO()
    bio.write("\n".join(lines).encode("utf-8"))
    bio.seek(0)
    return send_file(bio, mimetype="text/csv", as_attachment=True, download_name="inventaris.csv")

@app.get("/admin/inventory/export/xlsx")
def admin_inventory_export_xlsx():
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)
    
    with db() as conn:
        rows = conn.execute("SELECT * FROM inventory ORDER BY id DESC").fetchall()

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventaris Alkes"

    # 🏷️ Judul Header Kolom
    headers = [
        "Nama Alat", "Merk Alat", "Tipe Alat", "SN Alat", 
        "Ruangan", "Tahun Perolehan", "Sumber Dana",
        "Tanggal Kalibrasi", "Tanggal Jatuh Tempo", "Sertifikat Path"
    ]
    ws.append(headers)

    # 🎨 Styling Header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="12B886", end_color="12B886", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # 🧾 Tambahkan Data Inventaris
    for r in rows:
        ws.append([
            r["alat_nama"], r["alat_merk"], r["alat_tipe"], r["alat_sn"],
            r["ruangan"], r["tahun_perolehan"], r["sumber_dana"],
            r["tgl_kalibrasi"], r["tgl_jatuh_tempo"], r["sertifikat_path"] or ""
        ])

    # 📏 Styling seluruh data cell
    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

    # ⚙️ Auto Width tiap kolom
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # 🧱 Tinggi baris header
    ws.row_dimensions[1].height = 25

    # 💾 Simpan dan kirim file
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="inventaris_rapi.xlsx"
    )


# Tambah inventaris
@app.post("/admin/inventory/add")
def admin_inventory_add():
    if not session.get("user") or session["user"]["role"]!="admin":
        return abort(403)
    f = request.files.get("sertifikat")
    sert_path = None
    if f and f.filename:
        fname = secure_filename(f.filename)
        ext = os.path.splitext(fname)[1].lower()
        if ext not in ALLOWED_CERT_EXT:
            return "Format sertifikat harus PDF/JPG/PNG", 400
        newname = f"{int(datetime.now().timestamp())}_{fname}"
        # simpan file di folder static/uploads/certificates
        os.makedirs(UPLOAD_CERT, exist_ok=True)
        f.save(os.path.join(UPLOAD_CERT, newname))
        # simpan path relatif untuk url_for('static', filename=...)
        sert_path = f"uploads/certificates/{newname}"

    data = (
        request.form.get("alat_nama"),
        request.form.get("alat_merk"),
        request.form.get("alat_tipe"),
        request.form.get("alat_sn"),
        request.form.get("ruangan"),
        request.form.get("tahun_perolehan"),
        request.form.get("sumber_dana"),
        request.form.get("tgl_kalibrasi"),
        request.form.get("tgl_jatuh_tempo"),
        sert_path
    )
    with db() as conn:
        conn.execute("""INSERT INTO inventory(alat_nama,alat_merk,alat_tipe,alat_sn,ruangan,
                        tahun_perolehan,sumber_dana,tgl_kalibrasi,tgl_jatuh_tempo,sertifikat_path)
                        VALUES(?,?,?,?,?,?,?,?,?,?)""", data)
        conn.commit()
    return redirect(url_for("admin_inventory"))



# Halaman edit inventaris (GET)
@app.get("/admin/inventory/<int:iid>/edit")
def admin_inventory_edit(iid):
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)
    with db() as conn:
        item = conn.execute("SELECT * FROM inventory WHERE id=?", (iid,)).fetchone()
    if not item:
        return abort(404)

    # jika sertifikat tidak ada, beri nilai None
    sertifikat_url = url_for('static', filename=item['sertifikat_path']) if item['sertifikat_path'] else None

    return render_template("admin_inventory_edit.html",
                           item=item,
                           sertifikat_url=sertifikat_url,
                           BRAND=BRAND,
                           user=session.get("user"))

# Proses update inventaris (POST)
@app.post("/admin/inventory/<int:iid>/edit")
def admin_inventory_edit_post(iid):
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)
    f = request.files.get("sertifikat")
    with db() as conn:
        old_item = conn.execute("SELECT * FROM inventory WHERE id=?", (iid,)).fetchone()
        if not old_item:
            return abort(404)

        # jika ada file baru, simpan dan update path, kalau tidak pakai path lama
        if f and f.filename:
            fname = secure_filename(f.filename)
            ext = os.path.splitext(fname)[1].lower()
            if ext not in ALLOWED_CERT_EXT:
                return "Format sertifikat harus PDF/JPG/PNG", 400
            newname = f"{int(datetime.now().timestamp())}_{fname}"
            sert_path = os.path.join(UPLOAD_CERT, newname)
            f.save(sert_path)
            sert_path_rel = f"uploads/certificates/{newname}"  # pakai slash / dan relative path
        else:
            sert_path_rel = old_item["sertifikat_path"]

        # update database
        conn.execute("""UPDATE inventory SET
                        alat_nama=?,
                        alat_merk=?,
                        alat_tipe=?,
                        alat_sn=?,
                        ruangan=?,
                        tahun_perolehan=?,
                        sumber_dana=?,
                        tgl_kalibrasi=?,
                        tgl_jatuh_tempo=?,
                        sertifikat_path=?
                        WHERE id=?""",
                     (
                         request.form.get("alat_nama"),
                         request.form.get("alat_merk"),
                         request.form.get("alat_tipe"),
                         request.form.get("alat_sn"),
                         request.form.get("ruangan"),
                         request.form.get("tahun_perolehan"),
                         request.form.get("sumber_dana"),
                         request.form.get("tgl_kalibrasi"),
                         request.form.get("tgl_jatuh_tempo"),
                         sert_path_rel,
                         iid
                     ))
        conn.commit()

    return redirect(url_for("admin_inventory"))

# ---------- PEMELIHARAAN ----------
@app.get("/admin/maintenance")
def admin_maintenance():
    if not session.get("user") or session["user"]["role"] != "admin":
        return redirect(url_for("login"))

    tahun = request.args.get("tahun")

    # 🧠 Kalau tidak ada tahun di URL, ambil dari session
    if not tahun:
        tahun = session.get("tahun_aktif", datetime.now().year)
    else:
        # Simpan tahun yang dipilih user
        session["tahun_aktif"] = tahun

    q = (request.args.get("q") or "").strip()
    page = int(request.args.get("page", 1))
    per_page = 50
    offset = (page - 1) * per_page

    conn = sqlite3.connect("app.db")
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # 🔍 Filter data berdasarkan tahun dari semua kolom tanggal
    base_query = """
        FROM pemeliharaan
        WHERE (
            strftime('%Y', IFNULL(tgl_pemeliharaan, '0000-00-00')) = ?
            OR strftime('%Y', IFNULL(tgl_periode1, '0000-00-00')) = ?
            OR strftime('%Y', IFNULL(tgl_periode2, '0000-00-00')) = ?
            OR strftime('%Y', IFNULL(tgl_periode3, '0000-00-00')) = ?
        )
    """
    params = [str(tahun)] * 4  # ✅ INI HARUS DI DALAM BLOK, sejajar dengan base_query

    if q:
        base_query += """ AND (
            nama_alat LIKE ? OR merk LIKE ? OR tipe LIKE ? OR sn LIKE ? 
            OR ruangan LIKE ? OR kondisi LIKE ?
        )"""
        like = f"%{q}%"
        params.extend([like] * 6)

    # Hitung total data
    total_query = f"SELECT COUNT(*) {base_query}"
    cur.execute(total_query, params)
    total_records = cur.fetchone()[0]

    # Ambil data per halaman
    query = f"SELECT * {base_query} ORDER BY id DESC LIMIT ? OFFSET ?"
    params_with_pagination = params + [per_page, offset]
    cur.execute(query, params_with_pagination)
    records = cur.fetchall()

    total_pages = (total_records // per_page) + (1 if total_records % per_page else 0)
    conn.close()

    return render_template(
        "admin_maintenance.html",
        records=records,
        tahun=tahun,
        q=q,
        page=page,
        total_pages=total_pages,
        per_page=per_page
    )



# =========================================================
# 🔁 IMPORT DATA PEMELIHARAAN DARI TAHUN SEBELUMNYA
# =========================================================
@app.post("/import_maintenance_data/<int:tahun_asal>/<int:tahun_tujuan>")
def import_maintenance_data(tahun_asal, tahun_tujuan):
    try:
        with db() as conn:
            records = conn.execute("""
                SELECT nama_alat, merk, tipe, sn, ruangan, kondisi,
                       tgl_periode1, tgl_periode2, tgl_periode3,
                       file_periode1, file_periode2, file_periode3
                FROM pemeliharaan
                WHERE strftime('%Y', tgl_periode1) = ? 
                   OR strftime('%Y', tgl_periode2) = ? 
                   OR strftime('%Y', tgl_periode3) = ?
            """, (str(tahun_asal), str(tahun_asal), str(tahun_asal))).fetchall()

            print("=== IMPORT DIJALANKAN ===")
            print(f"Tahun Asal: {tahun_asal} → Tahun Tujuan: {tahun_tujuan}")
            print(f"Jumlah data ditemukan: {len(records)}")

            for r in records:
                print("Mengimpor:", r["nama_alat"])

                # 🔹 Ubah tahun pada tanggal lama ke tahun baru
                def ganti_tahun(tgl):
                    if not tgl:
                        return None
                    return tgl.replace(str(tahun_asal), str(tahun_tujuan))

                conn.execute("""
                    INSERT INTO pemeliharaan (
                        nama_alat, merk, tipe, sn, ruangan, kondisi,
                        tgl_periode1, tgl_periode2, tgl_periode3,
                        file_periode1, file_periode2, file_periode3
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    r["nama_alat"], r["merk"], r["tipe"], r["sn"], r["ruangan"], r["kondisi"],
                    ganti_tahun(r["tgl_periode1"]),
                    ganti_tahun(r["tgl_periode2"]),
                    ganti_tahun(r["tgl_periode3"]),
                    None, None, None  # kosongkan file untuk tahun baru
                ))

            conn.commit()
            flash(f"Berhasil mengimpor {len(records)} data dari tahun {tahun_asal} ke {tahun_tujuan}.", "success")

    except Exception as e:
        flash(f"Gagal mengimpor data: {e}", "danger")

    return redirect(url_for("admin_maintenance", tahun=tahun_tujuan))


# 🗑️ Hapus semua data berdasarkan tahun
@app.post("/admin/maintenance/delete_all/<int:tahun>")
def delete_all_maintenance(tahun):
    """Hapus semua data pemeliharaan berdasarkan tahun"""
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)

    with db() as conn:
        conn.execute("""
            DELETE FROM pemeliharaan
            WHERE strftime('%Y', tgl_pemeliharaan) = ? OR tgl_pemeliharaan IS NULL
        """, (str(tahun),))
        conn.commit()

    flash(f"🗑️ Semua data pemeliharaan tahun {tahun} telah dihapus.", "success")
    return redirect(url_for('admin_maintenance') + f'?tahun={tahun}')


# 📤 Export Excel per tahun (ambil tahun dari tanggal)
@app.route('/export_maintenance_excel')
def export_maintenance_excel():
    import sqlite3, os
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from flask import send_file, request

    tahun = request.args.get("tahun", datetime.now().year)
    conn = sqlite3.connect("app.db")
    c = conn.cursor()

    # 🧩 Cek dulu apakah tabel punya kolom "tahun"
    # kalau gak ada, deteksi tahun dari tanggal_periode (tahun pertama yg terisi)
    try:
        c.execute("""
            SELECT nama_alat, merk, tipe, sn, ruangan,
                   tgl_periode1, tgl_periode2, tgl_periode3,
                   kondisi,
                   file_periode1, file_periode2, file_periode3
            FROM pemeliharaan
            WHERE substr(tgl_periode1,1,4)=? 
               OR substr(tgl_periode2,1,4)=? 
               OR substr(tgl_periode3,1,4)=?
            ORDER BY id ASC
        """, (str(tahun), str(tahun), str(tahun)))
    except:
        c.execute("""
            SELECT nama_alat, merk, tipe, sn, ruangan,
                   tgl_periode1, tgl_periode2, tgl_periode3,
                   kondisi,
                   file_periode1, file_periode2, file_periode3
            FROM pemeliharaan
            ORDER BY id ASC
        """)
    
    data = c.fetchall()
    conn.close()

    # === Buat file Excel ===
    wb = Workbook()
    ws = wb.active
    ws.title = f"Pemeliharaan_{tahun}"

    headers = [
        "Nama Alat", "Merk", "Tipe", "SN", "Ruangan",
        "Periode 1", "Periode 2", "Periode 3",
        "Kondisi", "File P1", "File P2", "File P3"
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in data:
        ws.append([r if r else "" for r in row])

    # Auto width
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    # Simpan file
    export_dir = os.path.join("static", "exports")
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.join(export_dir, f"pemeliharaan_{tahun}.xlsx")
    wb.save(filepath)

    return send_file(filepath, as_attachment=True)




@app.post("/admin/maintenance/add")
def add_maintenance():
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)

    nama_alat = request.form["nama_alat"]
    merk = request.form["merk"]
    tipe = request.form["tipe"]
    sn = request.form["sn"]
    ruangan = request.form["ruangan"]
    kondisi = request.form["kondisi"]

    # Ambil tanggal tiap periode
    tgl1 = request.form.get("tgl_periode1")
    tgl2 = request.form.get("tgl_periode2")
    tgl3 = request.form.get("tgl_periode3")

    # Fungsi bantu simpan file
    def save_file(field):
        f = request.files.get(field)
        if f and f.filename:
            fname = secure_filename(f.filename)
            newname = f"{int(datetime.now().timestamp())}_{fname}"
            upload_folder = os.path.join(app.root_path, "static", "uploads", "pemeliharaan")
            os.makedirs(upload_folder, exist_ok=True)
            f.save(os.path.join(upload_folder, newname))
            return f"uploads/pemeliharaan/{newname}"
        return None

    # Simpan file untuk setiap periode
    file1 = save_file("file_periode1")
    file2 = save_file("file_periode2")
    file3 = save_file("file_periode3")

    with db() as conn:
        conn.execute("""
            INSERT INTO pemeliharaan
            (nama_alat, merk, tipe, sn, ruangan,
             tgl_periode1, tgl_periode2, tgl_periode3,
             kondisi, file_periode1, file_periode2, file_periode3)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (nama_alat, merk, tipe, sn, ruangan,
              tgl1, tgl2, tgl3, kondisi,
              file1, file2, file3))
        conn.commit()

    flash("✅ Data pemeliharaan berhasil ditambahkan dengan 3 periode.", "success")
    return redirect(url_for("admin_maintenance"))


@app.get("/admin/maintenance/<int:mid>/edit")
def admin_maintenance_edit(mid):
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)

    with db() as conn:
        record = conn.execute("SELECT * FROM pemeliharaan WHERE id=?", (mid,)).fetchone()
    if not record:
        return abort(404)

    return render_template(
        "admin_maintenance_edit.html",
        record=record,
        BRAND=BRAND,
        user=session.get("user")
    )

@app.post("/admin/maintenance/<int:mid>/edit")
def admin_maintenance_edit_post(mid):
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)

    nama_alat = request.form["nama_alat"]
    merk = request.form["merk"]
    tipe = request.form["tipe"]
    sn = request.form["sn"]
    ruangan = request.form["ruangan"]
    kondisi = request.form["kondisi"]

    # Ambil tanggal
    tgl1 = request.form.get("tgl_periode1")
    tgl2 = request.form.get("tgl_periode2")
    tgl3 = request.form.get("tgl_periode3")

    # Simpan file baru (kalau ada)
    def save_file(field):
        f = request.files.get(field)
        if f and f.filename:
            fname = secure_filename(f.filename)
            newname = f"{int(datetime.now().timestamp())}_{fname}"
            upload_folder = os.path.join(app.root_path, "static", "uploads", "pemeliharaan")
            os.makedirs(upload_folder, exist_ok=True)
            f.save(os.path.join(upload_folder, newname))
            return f"uploads/pemeliharaan/{newname}"
        return None

    file1 = save_file("file_periode1")
    file2 = save_file("file_periode2")
    file3 = save_file("file_periode3")

    with db() as conn:
        old = conn.execute("SELECT * FROM pemeliharaan WHERE id=?", (mid,)).fetchone()
        if not old:
            return abort(404)

        # kalau tidak upload baru, gunakan file lama
        if not file1:
            file1 = old["file_periode1"]
        if not file2:
            file2 = old["file_periode2"]
        if not file3:
            file3 = old["file_periode3"]

        conn.execute("""
            UPDATE pemeliharaan
            SET nama_alat=?, merk=?, tipe=?, sn=?, ruangan=?, 
                tgl_periode1=?, tgl_periode2=?, tgl_periode3=?,
                kondisi=?, file_periode1=?, file_periode2=?, file_periode3=?
            WHERE id=?
        """, (nama_alat, merk, tipe, sn, ruangan,
              tgl1, tgl2, tgl3,
              kondisi, file1, file2, file3, mid))
        conn.commit()

    flash("✅ Data pemeliharaan berhasil diperbarui.", "success")
    return redirect(url_for("admin_maintenance"))


@app.post("/admin/maintenance/<int:mid>/delete")
def admin_maintenance_delete(mid):
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)
    with db() as conn:
        row = conn.execute("SELECT file_path FROM pemeliharaan WHERE id=?", (mid,)).fetchone()
        if row and row['file_path']:
            try: os.remove(os.path.join(app.root_path, "static", row['file_path']))
            except: pass
        conn.execute("DELETE FROM pemeliharaan WHERE id=?", (mid,))
        conn.commit()
    flash("Data pemeliharaan berhasil dihapus.", "success")
    return redirect(url_for("admin_maintenance"))

# ---------- STOCK ----------
@app.get("/admin/stock/in")
def admin_stock_in():
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)
    
    page = int(request.args.get("page", 1))
    per_page = 50
    offset = (page - 1) * per_page

    with db() as conn:
        rows = conn.execute("""
            SELECT * FROM stock_in
            ORDER BY datetime(tgl_pembelian) DESC
            LIMIT ? OFFSET ?
        """, (per_page, offset)).fetchall()
        
        total_items = conn.execute("SELECT COUNT(*) as cnt FROM stock_in").fetchone()["cnt"]
        total_pages = (total_items + per_page - 1) // per_page

    return render_template(
        "admin_stock_in.html",
        items=rows,
        page=page,
        total_pages=total_pages,
        per_page=per_page,
        BRAND=BRAND,
        user=session.get("user")
    )


@app.post("/admin/stock/in/add")
def admin_stock_in_add():
    nama_barang = request.form["nama_barang"]
    qty = int(request.form["qty"])
    harga = float(request.form["harga"])
    tgl_pembelian = request.form["tgl_pembelian"]
    with db() as conn:
        conn.execute("INSERT INTO stock_in(nama_barang, qty, harga, tgl_pembelian) VALUES (?, ?, ?, ?)",
                     (nama_barang, qty, harga, tgl_pembelian))
        conn.commit()
    return redirect(url_for("admin_stock_in"))

# ---- Barang Keluar ----
@app.get("/admin/stock/out")
def admin_stock_out():
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)
    
    page = int(request.args.get("page", 1))
    per_page = 50
    offset = (page - 1) * per_page

    with db() as conn:
        rows = conn.execute("""
            SELECT * FROM stock_out
            ORDER BY datetime(tgl_keluar) DESC
            LIMIT ? OFFSET ?
        """, (per_page, offset)).fetchall()
        
        total_items = conn.execute("SELECT COUNT(*) as cnt FROM stock_out").fetchone()["cnt"]
        total_pages = (total_items + per_page - 1) // per_page

    return render_template(
        "admin_stock_out.html",
        items=rows,
        page=page,
        total_pages=total_pages,
        per_page=per_page,
        BRAND=BRAND,
        user=session.get("user")
    )



@app.post("/admin/stock/out/add")
def admin_stock_out_add():
    nama_barang = request.form["nama_barang"]
    qty = int(request.form["qty"])
    ruangan = request.form["ruangan"]
    penanggung_jawab = request.form["penanggung_jawab"]
    tgl_keluar = datetime.now().strftime("%Y-%m-%d")

    # Ambil file bukti (optional)
    bukti_file = request.files.get("bukti_file")
    bukti_path = None
    if bukti_file and bukti_file.filename:
        ext = os.path.splitext(bukti_file.filename)[1].lower()
        if ext in [".pdf", ".jpg", ".jpeg", ".png"]:
            upload_dir = os.path.join(app.root_path, "static", "uploads", "bukti_keluar")
            os.makedirs(upload_dir, exist_ok=True)
            filename = secure_filename(f"{int(datetime.now().timestamp())}_{bukti_file.filename}")
            bukti_file.save(os.path.join(upload_dir, filename))
            bukti_path = f"uploads/bukti_keluar/{filename}"

    with db() as conn:
        conn.execute(
            "INSERT INTO stock_out (nama_barang, qty, ruangan, penanggung_jawab, tgl_keluar, bukti_path) VALUES (?, ?, ?, ?, ?, ?)",
            (nama_barang, qty, ruangan, penanggung_jawab, tgl_keluar, bukti_path)
        )
        conn.commit()

    flash("Data barang keluar berhasil ditambahkan.", "success")
    return redirect(url_for("admin_stock_out"))


# ---- Rekapitulasi Stok ----
@app.get("/admin/stock/summary")
def admin_stock_summary():
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)

    q = (request.args.get("q") or "").strip()
    page = int(request.args.get("page", 1))
    per_page = 50
    offset = (page - 1) * per_page

    with db() as conn:
        if q:
            base_query = f"""
                SELECT nama_barang,
                       IFNULL((SELECT SUM(qty) FROM stock_in WHERE nama_barang = s.nama_barang), 0) -
                       IFNULL((SELECT SUM(qty) FROM stock_out WHERE nama_barang = s.nama_barang), 0)
                       AS qty_tersedia
                FROM (
                    SELECT nama_barang FROM stock_in
                    UNION
                    SELECT nama_barang FROM stock_out
                ) s
                WHERE nama_barang LIKE ?
                GROUP BY nama_barang
                ORDER BY nama_barang
                LIMIT ? OFFSET ?
            """
            rows = conn.execute(base_query, (f"%{q}%", per_page, offset)).fetchall()

            total_query = """
                SELECT COUNT(*) AS cnt FROM (
                    SELECT nama_barang FROM stock_in
                    UNION
                    SELECT nama_barang FROM stock_out
                ) s
                WHERE nama_barang LIKE ?
            """
            total_items = conn.execute(total_query, (f"%{q}%",)).fetchone()["cnt"]
        else:
            rows = conn.execute("""
                SELECT nama_barang,
                       IFNULL((SELECT SUM(qty) FROM stock_in WHERE nama_barang = s.nama_barang), 0) -
                       IFNULL((SELECT SUM(qty) FROM stock_out WHERE nama_barang = s.nama_barang), 0)
                       AS qty_tersedia
                FROM (
                    SELECT nama_barang FROM stock_in
                    UNION
                    SELECT nama_barang FROM stock_out
                ) s
                GROUP BY nama_barang
                ORDER BY nama_barang
                LIMIT ? OFFSET ?
            """, (per_page, offset)).fetchall()

            total_items = conn.execute("""
                SELECT COUNT(*) AS cnt FROM (
                    SELECT nama_barang FROM stock_in
                    UNION
                    SELECT nama_barang FROM stock_out
                ) s
            """).fetchone()["cnt"]

    total_pages = (total_items + per_page - 1) // per_page

    return render_template(
        "admin_stock_summary.html",
        items=rows,
        q=q,
        page=page,
        total_pages=total_pages,
        per_page=per_page,
        BRAND=BRAND,
        user=session.get("user")
    )



    return render_template(
        "admin_stock_summary.html",
        items=rows,
        q=q,  # kirim keyword ke template biar tetap muncul di input
        BRAND=BRAND,
        user=session.get("user")
    )

@app.get("/admin/stock/summary/export/xlsx")
def admin_stock_summary_export_xlsx():
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)

    q = (request.args.get("q") or "").strip()

    with db() as conn:
        if q:
            rows = conn.execute("""
                SELECT nama_barang,
                       IFNULL((SELECT SUM(qty) FROM stock_in WHERE nama_barang = s.nama_barang), 0) -
                       IFNULL((SELECT SUM(qty) FROM stock_out WHERE nama_barang = s.nama_barang), 0)
                       AS qty_tersedia
                FROM (
                    SELECT nama_barang FROM stock_in
                    UNION
                    SELECT nama_barang FROM stock_out
                ) s
                WHERE nama_barang LIKE ?
                GROUP BY nama_barang
                ORDER BY nama_barang
            """, (f"%{q}%",)).fetchall()
        else:
            rows = conn.execute("""
                SELECT nama_barang,
                       IFNULL((SELECT SUM(qty) FROM stock_in WHERE nama_barang = s.nama_barang), 0) -
                       IFNULL((SELECT SUM(qty) FROM stock_out WHERE nama_barang = s.nama_barang), 0)
                       AS qty_tersedia
                FROM (
                    SELECT nama_barang FROM stock_in
                    UNION
                    SELECT nama_barang FROM stock_out
                ) s
                GROUP BY nama_barang
                ORDER BY nama_barang
            """).fetchall()

    # Buat workbook
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Stok Barang Tersedia"

    # 🏷️ Header
    headers = ["Nama Barang", "Qty Tersedia"]
    ws.append(headers)

    # 🎨 Style header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="12B886", end_color="12B886", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # 🧾 Isi data
    for r in rows:
        ws.append([r["nama_barang"], r["qty_tersedia"]])

    # 🔧 Format seluruh tabel
    for row in ws.iter_rows(min_row=2, max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

    # 📏 Auto-width kolom
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # 🧱 Set tinggi header
    ws.row_dimensions[1].height = 25

    # 💾 Simpan ke buffer dan kirim
    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="stok_barang_tersedia.xlsx"
    )

# ---- Hapus Barang Masuk ----
@app.post("/admin/stock/in/<int:item_id>/delete")
def admin_stock_in_delete(item_id):
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)
    with db() as conn:
        conn.execute("DELETE FROM stock_in WHERE id=?", (item_id,))
        conn.commit()
    flash("Data barang masuk berhasil dihapus.", "success")
    return redirect(url_for("admin_stock_in"))

# ---- Hapus Barang Keluar ----
@app.post("/admin/stock/out/<int:item_id>/delete")
def admin_stock_out_delete(item_id):
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)
    with db() as conn:
        conn.execute("DELETE FROM stock_out WHERE id=?", (item_id,))
        conn.commit()
    flash("Data barang keluar berhasil dihapus.", "success")
    return redirect(url_for("admin_stock_out"))
    
@app.post("/admin/stock/out/<int:item_id>/upload")
def admin_stock_out_upload(item_id):
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)

    file = request.files.get("bukti_file")
    if not file or not file.filename:
        flash("Tidak ada file yang diunggah.", "warning")
        return redirect(url_for("admin_stock_out"))

    # Validasi ekstensi
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in [".pdf", ".jpg", ".jpeg", ".png"]:
        flash("Format file tidak diizinkan (hanya PDF/JPG/PNG).", "danger")
        return redirect(url_for("admin_stock_out"))

    # Simpan file
    upload_dir = os.path.join(app.root_path, "static", "uploads", "bukti_keluar")
    os.makedirs(upload_dir, exist_ok=True)
    filename = secure_filename(f"{int(datetime.now().timestamp())}_{file.filename}")
    file.save(os.path.join(upload_dir, filename))

    rel_path = f"uploads/bukti_keluar/{filename}"

    with db() as conn:
        conn.execute("UPDATE stock_out SET bukti_path=? WHERE id=?", (rel_path, item_id))
        conn.commit()

    flash("Bukti barang keluar berhasil diunggah.", "success")
    return redirect(url_for("admin_stock_out"))
    
@app.post("/admin/stock/out/<int:item_id>/edit-file")
def admin_stock_out_edit_file(item_id):
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)

    file = request.files.get("bukti_file")
    if not file or not file.filename:
        flash("Tidak ada file yang diunggah.", "warning")
        return redirect(url_for("admin_stock_out"))

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in [".pdf", ".jpg", ".jpeg", ".png"]:
        flash("Format file tidak diizinkan (hanya PDF/JPG/PNG).", "danger")
        return redirect(url_for("admin_stock_out"))

    upload_dir = os.path.join(app.root_path, "static", "uploads", "bukti_keluar")
    os.makedirs(upload_dir, exist_ok=True)
    filename = secure_filename(f"{int(datetime.now().timestamp())}_{file.filename}")
    file.save(os.path.join(upload_dir, filename))

    rel_path = f"uploads/bukti_keluar/{filename}"

    with db() as conn:
        conn.execute("UPDATE stock_out SET bukti_path=? WHERE id=?", (rel_path, item_id))
        conn.commit()

    flash("Bukti barang keluar berhasil diperbarui.", "success")
    return redirect(url_for("admin_stock_out"))
    



# ---------- SPO (ADMIN & USER) ----------
@app.get("/admin/spo")
def admin_spo():
    if not session.get("user") or session["user"]["role"]!="admin":
        return abort(403)

    with db() as conn:
        spo_items = conn.execute("SELECT * FROM spo ORDER BY id DESC").fetchall()

    return render_template("admin_spo.html", spo_items=spo_items,
                           BRAND=BRAND, user=session.get("user"))


@app.post("/admin/spo/add")
def admin_spo_add():
    if not session.get("user") or session["user"]["role"]!="admin":
        return abort(403)

    f = request.files.get("spo_file")
    file_path = None
    if f and f.filename:
        fname = secure_filename(f.filename)
        newname = f"{int(datetime.now().timestamp())}_{fname}"
        upload_folder = os.path.join("static", "uploads", "spo")
        os.makedirs(upload_folder, exist_ok=True)
        f.save(os.path.join(upload_folder, newname))
        file_path = f"uploads/spo/{newname}"

    data = (
        request.form.get("alat_nama"),
        request.form.get("alat_merk"),
        request.form.get("alat_tipe"),
        file_path
    )
    with db() as conn:
        conn.execute("INSERT INTO spo(alat_nama,alat_merk,alat_tipe,file_path) VALUES(?,?,?,?)", data)
        conn.commit()

    return redirect(url_for("admin_spo"))

# Halaman edit SPO (GET)
@app.get("/admin/spo/<int:sid>/edit")
def admin_spo_edit(sid):
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)

    with db() as conn:
        spo_item = conn.execute("SELECT * FROM spo WHERE id=?", (sid,)).fetchone()
    if not spo_item:
        return abort(404)

    return render_template("admin_spo_edit.html",
                           spo_item=spo_item,
                           BRAND=BRAND,
                           user=session.get("user"))

# Proses update SPO (POST)
@app.post("/admin/spo/<int:sid>/edit")
def admin_spo_edit_post(sid):
    if not session.get("user") or session["user"]["role"] != "admin":
        return abort(403)

    f = request.files.get("spo_file")
    with db() as conn:
        old_item = conn.execute("SELECT * FROM spo WHERE id=?", (sid,)).fetchone()
        if not old_item:
            return abort(404)

        file_path = old_item["file_path"]
        if f and f.filename:
            fname = secure_filename(f.filename)
            newname = f"{int(datetime.now().timestamp())}_{fname}"
            upload_folder = os.path.join("static", "uploads", "spo")
            os.makedirs(upload_folder, exist_ok=True)
            f.save(os.path.join(upload_folder, newname))
            file_path = f"uploads/spo/{newname}"

        conn.execute("""UPDATE spo SET alat_nama=?, alat_merk=?, alat_tipe=?, file_path=? WHERE id=?""",
                     (request.form.get("alat_nama"),
                      request.form.get("alat_merk"),
                      request.form.get("alat_tipe"),
                      file_path,
                      sid))
        conn.commit()
    return redirect(url_for("admin_spo"))

@app.post("/admin/spo/<int:sid>/delete")
def admin_spo_delete(sid):
    if not session.get("user") or session["user"]["role"]!="admin":
        return abort(403)
    with db() as conn:
        row = conn.execute("SELECT file_path FROM spo WHERE id=?", (sid,)).fetchone()
        if row and row["file_path"]:
            try: os.remove(os.path.join(BASE_DIR, "static", row["file_path"]))
            except: pass
        conn.execute("DELETE FROM spo WHERE id=?", (sid,))
        conn.commit()
    return redirect(url_for("admin_spo"))




# ---------- USER ----------
@app.get("/user")
def user_dashboard():
    if not session.get("user") or session["user"]["role"] != "user":
        return abort(403)

    page = int(request.args.get("page", 1))
    per_page = 50
    offset = (page - 1) * per_page

    with db() as conn:
        rows = conn.execute("""
            SELECT * FROM reports
            WHERE user_id=?
            ORDER BY datetime(created_at) DESC
            LIMIT ? OFFSET ?
        """, (session["user"]["id"], per_page, offset)).fetchall()

        total_reports = conn.execute("""
            SELECT COUNT(*) as cnt FROM reports WHERE user_id=?
        """, (session["user"]["id"],)).fetchone()["cnt"]

        total_pages = (total_reports + per_page - 1) // per_page

    return render_template(
        "user_dashboard.html",
        reports=rows,
        page=page,
        total_pages=total_pages,
        per_page=per_page,
        BRAND=BRAND,
        user=session.get("user")
    )



@app.get("/user/report/new")
def user_report_new():
    if not session.get("user") or session["user"]["role"]!="user":
        return abort(403)
    return render_template("user_report_new.html", BRAND=BRAND, user=session.get("user"))

@app.post("/user/report/new")
def user_report_new_post():
    if not session.get("user") or session["user"]["role"] != "user":
        return abort(403)

    title = request.form.get("title")
    unit = request.form.get("unit")
    description = request.form.get("description")
    img = request.files.get("image")
    image_path = None

    if img and img.filename:
        fname = secure_filename(img.filename)
        timestamp = int(datetime.now().timestamp())
        newname = f"{timestamp}_{fname}"
        upload_folder_abs = os.path.join(BASE_DIR, "static", "uploads", "reports")
        os.makedirs(upload_folder_abs, exist_ok=True)
        img.save(os.path.join(upload_folder_abs, newname))
        image_path = f"uploads/reports/{newname}"

    tz = pytz.timezone("Asia/Jakarta")
    created_at = datetime.now(tz).strftime("%Y-%m-%d %H:%M:%S")

    with db() as conn:
        conn.execute(
            """INSERT INTO reports(title,unit,description,image_path,status,created_at,user_id)
               VALUES(?,?,?,?,?,?,?)""",
            (title, unit, description, image_path, "Proses", created_at, session["user"]["id"])
        )
        conn.commit()
        rid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]

    # 🔹 Emit realtime event hanya ke admin
    new_report_data = {
        "id": rid,
        "title": title,
        "unit": unit,
        "description": description,
        "image_path": image_path,
        "status": "Proses",
        "comment": "",
        "created_at": created_at
    }
    socketio.emit("new_report", new_report_data, room="admin")
    print(f"📢 Emit new_report -> room=admin : {new_report_data}")  # <-- 🔍 Tambahan log debug

    return redirect(url_for("user_dashboard"))


# ---------- USER SPO ----------
@app.get("/user/spo")
def user_spo():
    if not session.get("user") or session["user"]["role"] != "user":
        return abort(403)

    with db() as conn:
        spo_items = conn.execute("SELECT * FROM spo ORDER BY id DESC").fetchall()

    return render_template("user_spo.html",
                           spo_items=spo_items,
                           BRAND=BRAND,
                           user=session.get("user"))


def init():
    with db() as conn:
        pass
    ensure_tables()
    ensure_admin()

@app.get("/chat")
def chat_page():
    if not session.get("user"):
        return redirect(url_for("login"))

    with db() as conn:
        users = conn.execute(
            "SELECT id, username, role FROM users WHERE id != ?",
            (session["user"]["id"],)
        ).fetchall()
        messages = conn.execute("""
            SELECT m.*, u.username as sender_name 
            FROM messages m 
            JOIN users u ON m.sender_id = u.id
            ORDER BY datetime(m.created_at) ASC
        """).fetchall()

    return render_template("chat.html",
                           users=users,
                           messages=messages,
                           user=session["user"],
                           BRAND=BRAND)

# === SOCKETIO EVENTS ===

# Simpan daftar user yang sudah join agar tidak spam notifikasi
joined_users = set()

@socketio.on('join')
def handle_join(data):
    username = data.get('username')
    room = data.get('room', 'umum')

    join_room(room)
    print(f"🟢 {username} bergabung ke room {room}")

    # ✅ hanya kirim notifikasi sekali per sesi (tidak setiap aksi / reload)
    if username not in joined_users:
        emit('message', {
            'user': 'System',
            'msg': f'{username} bergabung ke ruang {room}'
        }, room=room)
        joined_users.add(username)



@socketio.on('send_message')
def handle_send_message(data):
    username = data.get('username')
    room = data.get('room', 'umum')
    msg = data.get('msg')

    # --- kirim pesan realtime ke semua pengguna di room itu ---
    emit('message', {'user': username, 'msg': msg}, room=room)

    # --- kirim notifikasi ke semua yang join room (kecuali pengirim) ---
    emit('chat_notification', {'from': username, 'msg': msg}, room=room, include_self=False)

    # --- 🧩 simpan juga ke database messages ---
    sender_id = session['user']['id']
    receiver_id = data.get('receiver_id', 1)  # contoh: admin = 1

    import pytz
    tz = pytz.timezone("Asia/Jakarta")
    created_at = datetime.now(tz).strftime("%Y-%m-%d %H:%M:%S")

    with db() as conn:
        conn.execute("""
            INSERT INTO messages(sender_id, receiver_id, message, created_at)
            VALUES (?, ?, ?, ?)
        """, (sender_id, receiver_id, msg, created_at))
        conn.commit()




if __name__ == "__main__":
    import sys
    init()
    if "--init-db" in sys.argv:
        print("Database siap.")
        raise SystemExit(0)

    # --- DETEKSI IP LAN DAN CETAK URL ---
    import socket

    def get_local_ip():
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
        except:
            ip = "127.0.0.1"
        finally:
            s.close()
        return ip

    local_ip = get_local_ip()
    print("========================================")
    print("Flask server siap dijalankan!")
    print("Akses dari komputer ini: http://localhost:5000")
    print(f"Akses dari jaringan LAN: http://{local_ip}:5000")
    print("Catatan: untuk akses publik, lakukan port forwarding dan nonaktifkan debug=True")
    print("========================================")

    socketio.run(app, host="0.0.0.0", port=5000, debug=True)

